import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import random


#function to download prices from ebay by being passed a keyword to search
#it is important to pass UPC as the keyword to prevent other results from showing
#if there are a limited number of results the scraper is not smart enought to distinguish between results and other suggested items 


ship_price_wb = openpyxl.load_workbook('shipping.xlsx')
ship_price_ws = ship_price_wb.active
ship_price_tuple = tuple(ship_price_ws.columns)
ship_price_dictionary = dict()
for i in range(len(ship_price_tuple[0])):
    ship_price_dictionary[ship_price_tuple[0][i].value] = ship_price_tuple[1][i].value


def ebay_price_lookup(ebay_url):
	ebay_competitor_pricing = []
	#sets the ebay url with query selectors
	#&LH_BIN=1  requires buy it now items
	#LH_FS=1    requires free shipping (workaround:couldn't figure out how to scrape shipping price)
	#&_sop=15   price + shipping: low to high
	ebay_url_prepend = 'https://www.ebay.com/sch/i.html?LH_BIN=1&LH_FS=1&_sop=15&_nkw='
	
    #pulls the html page and parses it into a readable format
	r = requests.get(ebay_url_prepend + str(ebay_url))
	ebay_page_soup = BeautifulSoup(r.text, 'html.parser')
	#search for all bold tags (only price has a span w/ class "bold")
	price_list = ebay_page_soup.find_all('span', class_= 'bold')
	for i in price_list:
		#replace filler html
		ebay_competitor_pricing.append(float(i.text.replace('\n', '').replace('\t', '').replace('$', '')))
	return (ebay_competitor_pricing)

#read excel file with list of SKU UPC's in column A of the Active Sheet. Please keep excel columns as is
#the following will pull the data from excel and query ebay for the provided UPCs

excel_workbook = openpyxl.load_workbook('testtest.xlsx')
excel_sheet = excel_workbook.active
excel_columns = tuple(excel_sheet.columns)

UPC_list = excel_sheet['B']
List_Price = excel_sheet['C']
Cost_list = excel_sheet['E']
Weight_list = excel_sheet['F']
Dimension1_list = excel_sheet['G']
Dimension2_list = excel_sheet['H']
Dimension3_list = excel_sheet['I']
price_lookup=['List_Price']
profit_list = ['Expected Profit']

backup_price_wb = openpyxl.Workbook()
backup_price_ws = backup_price_wb.active
backup_price_ws.cell(column=1, row=1).value = 'UPC'
backup_price_ws.cell(column=2, row=1).value = 'Low Price 1'
backup_price_ws.cell(column=3, row=1).value = 'Low Price 2'
backup_price_ws.cell(column=4, row=1).value = 'Low Price 3... etc'

for i in range(1, len(UPC_list)):
    price_has_been_removed = False
    competitor_below_breakeven = False
    time.sleep(random.randint(5,20)/10)
    previous_list_price = List_Price[i].value
    dimensions = sorted([Dimension1_list[i].value, Dimension2_list[i].value, Dimension3_list[i].value])
    package_volume = dimensions[0]*dimensions[1]*dimensions[2]
    #need to make sure that there is a dimension for every part (no zeros allowed)
    #also they all need to be ints or floats
    if package_volume <= 1728:
        dimensional_weight_denomenator = 166
    else:
        dimensional_weight_denomenator = 139
    dimensional_weight = round(package_volume/dimensional_weight_denomenator)
    large_package_check = 2*dimensions[0]+2*dimensions[1]*2 + dimensions[2]
    if large_package_check > 130 and dimensional_weight < 90:
        dimensional_weight = 90
    billable_weight = round(max(dimensional_weight, Weight_list[i].value))
    ship_cost = ship_price_dictionary[billable_weight]
    cost_of_goods = Cost_list[i].value
    estimated_break_even_price = (cost_of_goods+ship_cost)/.9
    #need to handle the case where a cell is blank (causes error when performing math on ebay_competitor_pricing)
    if UPC_list[i].value is not None:
        ebay_competitor_pricing = ebay_price_lookup(UPC_list[i].value)
        try:
            while ebay_competitor_pricing[0] < estimated_break_even_price:
                del ebay_competitor_pricing[0]
        except:
            competitor_below_breakeven = True
        try:
            ebay_competitor_pricing.remove(previous_list_price)
            #handle error if the item has QTY=0 and isn't active on ebay
            price_has_been_removed = True
        except:
            pass
    else:
        ebay_competitor_pricing(['error in url'])
    #used a weighted mean to calculate list price
    try:
        our_new_list_price = (.25*ebay_competitor_pricing[0]+.65*ebay_competitor_pricing[1]+.1*ebay_competitor_pricing[2])
        notes = '3+ listings found (excluding us). Price set to .25*First + .65*Second +.1*Third'
    except TypeError:
        our_new_list_price = estimated_break_even_price*1.1
        notes = 'all competitor pricing below breakeven. Price set at Breakeven times 10%'
    except IndexError:
        try:
            our_new_list_price = (.3*ebay_competitor_pricing[0]+.7*ebay_competitor_pricing[1])
            notes = '2 listings found (excluding us). Price set to .3*First + .7*Second'
        except IndexError:
            try:
                our_new_list_price = ebay_competitor_pricing[0]-.05
                notes = 'One listing found (excluding us). Price set too .05$ less than them'
            except IndexError:
                if competitor_below_breakeven == False and price_has_been_removed == False:
                    our_new_list_price = '=#N/A'
                    notes = 'no data pulled'
                elif competitor_below_breakeven == False and price_has_been_removed == True:
                    our_new_list_price = previous_list_price*1.1
                    notes = 'We are the only lister. Price raised 10%'
                else:
                    our_new_list_price = estimated_break_even_price*1.1
                    notes = 'All competitors are below breakeven. Price set at 110% of breakeven'
    #can probably remove the cost lookup from python no need for it
    if type(our_new_list_price) == float:
        our_new_list_price = round(our_new_list_price, 2)
    expected_profit = "=.9*C"+str(i+1)+"-E"+str(i+1)+"-"+str(ship_cost)
    excel_sheet.cell(column=3, row=i+1).value = our_new_list_price
    excel_sheet.cell(column=4, row=i+1).value = expected_profit
    excel_sheet.cell(column=10, row=i+1).value = notes
    backup_price_ws.cell(column=1, row=i+1).value = UPC_list[i].value
    for j in range(len(ebay_competitor_pricing)):
        backup_price_ws.cell(column=j+2, row=i+1).value = ebay_competitor_pricing[j]

backup_price_wb.save(filename='testtestprices.xlsx')
excel_workbook.save(filename='testtestresults.xlsx')